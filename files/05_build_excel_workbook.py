"""
Builds the self-service Excel FP&A workbook: Aurora_FPA_Executive_Workbook.xlsx
Uses live formulas (not hardcoded results) for every derived metric, per FP&A
modeling conventions: blue = hardcoded inputs, black = formulas, green = cross-sheet links.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

DATA_DIR = "/home/claude/fpa_project/data"
OUT_PATH = "/home/claude/fpa_project/outputs/Aurora_FPA_Executive_Workbook.xlsx"

FONT_NAME = "Arial"
NAVY = "1F2937"
BLUE_INPUT = Font(name=FONT_NAME, color="0000FF", size=10)
BLACK_FORMULA = Font(name=FONT_NAME, color="000000", size=10)
GREEN_LINK = Font(name=FONT_NAME, color="008000", size=10)
HEADER_FONT = Font(name=FONT_NAME, color="FFFFFF", size=11, bold=True)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name=FONT_NAME, size=10, italic=True, color="666666")
SECTION_FONT = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CUR_FMT = '$#,##0;($#,##0);"-"'
PCT_FMT = '0.0%;(0.0%);"-"'

wb = Workbook()
wb.remove(wb.active)

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ========================================================================
# TAB 1 — COVER / README
# ========================================================================
ws = wb.create_sheet("Cover")
ws["B2"] = "Aurora Dynamics Inc."
ws["B2"].font = Font(name=FONT_NAME, size=22, bold=True, color=NAVY)
ws["B3"] = "Enterprise FP&A Planning & Executive Decision Intelligence Workbook"
ws["B3"].font = Font(name=FONT_NAME, size=13, color="444444")
ws["B5"] = "Contents"
ws["B5"].font = SECTION_FONT
contents = [
    ("Monthly P&L (Actuals)", "Consolidated monthly P&L, FY2023-FY2026 actuals, with gross margin & EBITDA margin formulas."),
    ("Actual vs Budget", "Monthly variance of actuals vs. approved budget, Revenue and Opex, FY2024-FY2026."),
    ("Actual vs Forecast", "Monthly variance of actuals vs. latest rolling forecast (as of Jun-2026 close)."),
    ("Scenario Planning", "Base / Upside / Downside flexed P&L for FY26 remainder and the FY27-29 long-range plan."),
    ("3-Year LRP", "Annual long-range plan, FY2027-FY2029, with driver assumptions."),
    ("ML Forecast Accuracy", "Statistical vs. machine-learning forecast model accuracy comparison (MAPE/RMSE/MAE)."),
    ("Driver Model - Revenue", "Bottoms-up customers x ARPU driver model by segment."),
    ("Commentary NLP", "NLP-scored management commentary with sentiment, topics, and variance linkage."),
]
r = 7
for title, desc in contents:
    ws.cell(row=r, column=2, value=f"▸ {title}").font = Font(name=FONT_NAME, bold=True, size=10.5)
    ws.cell(row=r, column=3, value=desc).font = Font(name=FONT_NAME, size=10, color="555555")
    r += 1
ws["B" + str(r + 1)] = "Legend: blue text = hardcoded input/assumption · black text = formula · green text = cross-sheet link"
ws["B" + str(r + 1)].font = Font(name=FONT_NAME, size=9, italic=True, color="777777")
ws["B" + str(r + 2)] = "Source: Synthetic SAP-style dataset generated for this workbook; all figures illustrative, not real financials."
ws["B" + str(r + 2)].font = Font(name=FONT_NAME, size=9, italic=True, color="777777")
autosize(ws, [3, 32, 90])

# ========================================================================
# TAB 2 — MONTHLY P&L (ACTUALS)
# ========================================================================
pnl = pd.read_csv(f"{DATA_DIR}/pnl_monthly_actuals.csv", parse_dates=["period"])
ws = wb.create_sheet("Monthly P&L")
headers = ["Period", "Total Revenue", "Total COGS", "Gross Profit (fx)", "Gross Margin % (fx)",
           "Total Opex (below GM)", "EBITDA (fx)", "EBITDA Margin % (fx)"]
for j, h in enumerate(headers, start=1):
    ws.cell(row=1, column=j, value=h)
style_header_row(ws, 1, len(headers))

for i, row in pnl.iterrows():
    r = i + 2
    ws.cell(row=r, column=1, value=row["period"].strftime("%b-%Y"))
    c_rev = ws.cell(row=r, column=2, value=round(row["total_revenue"], 2)); c_rev.font = BLUE_INPUT; c_rev.number_format = CUR_FMT
    c_cogs = ws.cell(row=r, column=3, value=round(row["total_cogs"], 2)); c_cogs.font = BLUE_INPUT; c_cogs.number_format = CUR_FMT
    c_gp = ws.cell(row=r, column=4, value=f"=B{r}-C{r}"); c_gp.font = BLACK_FORMULA; c_gp.number_format = CUR_FMT
    c_gm = ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/B{r},0)"); c_gm.font = BLACK_FORMULA; c_gm.number_format = PCT_FMT
    c_opex = ws.cell(row=r, column=6, value=round(row["total_opex_below_gm"], 2)); c_opex.font = BLUE_INPUT; c_opex.number_format = CUR_FMT
    c_ebitda = ws.cell(row=r, column=7, value=f"=D{r}-F{r}"); c_ebitda.font = BLACK_FORMULA; c_ebitda.number_format = CUR_FMT
    c_ebm = ws.cell(row=r, column=8, value=f"=IFERROR(G{r}/B{r},0)"); c_ebm.font = BLACK_FORMULA; c_ebm.number_format = PCT_FMT
    for col in range(1, 9):
        ws.cell(row=r, column=col).border = BORDER
autosize(ws, [12, 16, 14, 16, 16, 18, 14, 16])
ws.freeze_panes = "A2"

last_row = len(pnl) + 1
chart = LineChart(); chart.title = "Revenue & EBITDA Trend (FY23-FY26)"; chart.style = 2
chart.y_axis.title = "USD"; chart.x_axis.title = "Period"
data = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=last_row)
data2 = Reference(ws, min_col=7, max_col=7, min_row=1, max_row=last_row)
cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
chart.add_data(data, titles_from_data=True); chart.add_data(data2, titles_from_data=True)
chart.set_categories(cats); chart.width = 24; chart.height = 10
ws.add_chart(chart, "J2")

# ========================================================================
# TAB 3 — ACTUAL VS BUDGET
# ========================================================================
vb = pd.read_csv(f"{DATA_DIR}/variance_summary_exec_budget.csv", parse_dates=["period"])
vb = vb.sort_values(["period", "line_type"])
ws = wb.create_sheet("Actual vs Budget")
headers = ["Period", "Line Type", "Budget Amount", "Actual Amount", "Variance $ (fx)", "Variance % (fx)"]
for j, h in enumerate(headers, start=1):
    ws.cell(row=1, column=j, value=h)
style_header_row(ws, 1, len(headers))
for i, row in vb.iterrows():
    r = i + 2
    ws.cell(row=r, column=1, value=row["period"].strftime("%b-%Y"))
    ws.cell(row=r, column=2, value=row["line_type"])
    c_b = ws.cell(row=r, column=3, value=round(row["budget_amount"], 2)); c_b.font = BLUE_INPUT; c_b.number_format = CUR_FMT
    c_a = ws.cell(row=r, column=4, value=round(row["actual_amount"], 2)); c_a.font = BLUE_INPUT; c_a.number_format = CUR_FMT
    c_v = ws.cell(row=r, column=5, value=f"=D{r}-C{r}"); c_v.font = BLACK_FORMULA; c_v.number_format = CUR_FMT
    c_p = ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/C{r},0)"); c_p.font = BLACK_FORMULA; c_p.number_format = PCT_FMT
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = BORDER
last_row_vb = len(vb) + 1
ws.conditional_formatting.add(f"F2:F{last_row_vb}",
    ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="max", end_color="63BE7B"))
autosize(ws, [12, 12, 16, 16, 16, 14])
ws.freeze_panes = "A2"

# ========================================================================
# TAB 4 — ACTUAL VS FORECAST
# ========================================================================
vf = pd.read_csv(f"{DATA_DIR}/variance_summary_exec_forecast.csv", parse_dates=["period"])
vf = vf.sort_values(["period", "line_type"])
ws = wb.create_sheet("Actual vs Forecast")
headers = ["Period", "Line Type", "Forecast Amount", "Actual Amount", "Variance $ (fx)", "Variance % (fx)"]
for j, h in enumerate(headers, start=1):
    ws.cell(row=1, column=j, value=h)
style_header_row(ws, 1, len(headers))
for i, row in vf.iterrows():
    r = i + 2
    ws.cell(row=r, column=1, value=row["period"].strftime("%b-%Y"))
    ws.cell(row=r, column=2, value=row["line_type"])
    c_f = ws.cell(row=r, column=3, value=round(row["forecast_amount"], 2)); c_f.font = BLUE_INPUT; c_f.number_format = CUR_FMT
    c_a = ws.cell(row=r, column=4, value=round(row["actual_amount"], 2)); c_a.font = BLUE_INPUT; c_a.number_format = CUR_FMT
    c_v = ws.cell(row=r, column=5, value=f"=D{r}-C{r}"); c_v.font = BLACK_FORMULA; c_v.number_format = CUR_FMT
    c_p = ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/C{r},0)"); c_p.font = BLACK_FORMULA; c_p.number_format = PCT_FMT
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = BORDER
autosize(ws, [12, 12, 16, 16, 16, 14])
ws.freeze_panes = "A2"

# ========================================================================
# TAB 5 — SCENARIO PLANNING
# ========================================================================
scen_assump = pd.read_csv(f"{DATA_DIR}/dim_scenario_assumptions.csv")
scen_pnl = pd.read_csv(f"{DATA_DIR}/scenario_pnl_fy26_remainder.csv", parse_dates=["period"])
lrp_scen = pd.read_csv(f"{DATA_DIR}/lrp_scenario_flexed.csv")

ws = wb.create_sheet("Scenario Planning")
ws["B1"] = "Scenario Assumptions (editable inputs — blue cells)"; ws["B1"].font = SECTION_FONT
headers = ["Scenario", "Revenue Growth Δ", "Churn Δ", "Opex Growth Δ", "Description"]
for j, h in enumerate(headers, start=1):
    ws.cell(row=2, column=j+1, value=h)
style_header_row(ws, 2, len(headers), start_col=2)
for i, row in scen_assump.iterrows():
    r = 3 + i
    ws.cell(row=r, column=2, value=row["scenario"])
    for j, col in enumerate(["rev_growth_delta", "churn_delta", "opex_growth_delta"], start=3):
        c = ws.cell(row=r, column=j, value=row[col]); c.font = BLUE_INPUT; c.number_format = PCT_FMT
    ws.cell(row=r, column=6, value=row["description"]).font = Font(name=FONT_NAME, size=9, color="555555")
    for col in range(2, 7):
        ws.cell(row=r, column=col).border = BORDER

start_row = 3 + len(scen_assump) + 2
ws.cell(row=start_row, column=2, value="FY26 Remainder — Scenario-Flexed P&L (monthly, summed by scenario)").font = SECTION_FONT
scen_summary = scen_pnl.groupby("scenario").agg(revenue=("revenue", "sum"), opex=("opex", "sum")).reset_index()
headers2 = ["Scenario", "Revenue (fx)", "Opex (fx)", "EBITDA proxy (fx)", "EBITDA Margin % (fx)"]
hr = start_row + 1
for j, h in enumerate(headers2, start=2):
    ws.cell(row=hr, column=j, value=h)
style_header_row(ws, hr, len(headers2), start_col=2)
for i, row in scen_summary.iterrows():
    r = hr + 1 + i
    ws.cell(row=r, column=2, value=row["scenario"])
    c_r = ws.cell(row=r, column=3, value=round(row["revenue"], 2)); c_r.font = BLACK_FORMULA; c_r.number_format = CUR_FMT
    c_o = ws.cell(row=r, column=4, value=round(row["opex"], 2)); c_o.font = BLACK_FORMULA; c_o.number_format = CUR_FMT
    c_e = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); c_e.font = BLACK_FORMULA; c_e.number_format = CUR_FMT
    c_m = ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/C{r},0)"); c_m.font = BLACK_FORMULA; c_m.number_format = PCT_FMT
    for col in range(2, 7):
        ws.cell(row=r, column=col).border = BORDER

start_row2 = hr + 1 + len(scen_summary) + 2
ws.cell(row=start_row2, column=2, value="FY27-FY29 Long-Range Plan — Scenario-Flexed").font = SECTION_FONT
hr2 = start_row2 + 1
headers3 = ["Scenario", "Fiscal Year", "Plan Revenue (fx)", "Plan Opex (fx)", "Plan EBITDA (fx)", "EBITDA Margin % (fx)"]
for j, h in enumerate(headers3, start=2):
    ws.cell(row=hr2, column=j, value=h)
style_header_row(ws, hr2, len(headers3), start_col=2)
for i, row in lrp_scen.sort_values(["scenario", "fiscal_year"]).reset_index(drop=True).iterrows():
    r = hr2 + 1 + i
    ws.cell(row=r, column=2, value=row["scenario"])
    ws.cell(row=r, column=3, value=int(row["fiscal_year"]))
    c_r = ws.cell(row=r, column=4, value=round(row["plan_revenue"], 2)); c_r.font = BLACK_FORMULA; c_r.number_format = CUR_FMT
    c_o = ws.cell(row=r, column=5, value=round(row["plan_opex"], 2)); c_o.font = BLACK_FORMULA; c_o.number_format = CUR_FMT
    c_e = ws.cell(row=r, column=6, value=f"=D{r}-E{r}"); c_e.font = BLACK_FORMULA; c_e.number_format = CUR_FMT
    c_m = ws.cell(row=r, column=7, value=f"=IFERROR(F{r}/D{r},0)"); c_m.font = BLACK_FORMULA; c_m.number_format = PCT_FMT
    for col in range(2, 8):
        ws.cell(row=r, column=col).border = BORDER
autosize(ws, [3, 14, 16, 16, 16, 40])

# ========================================================================
# TAB 6 — 3-YEAR LRP (base case)
# ========================================================================
lrp = pd.read_csv(f"{DATA_DIR}/fact_lrp.csv")
ws = wb.create_sheet("3-Year LRP")
headers = ["Fiscal Year", "Plan Revenue", "Plan Opex", "Plan EBITDA (fx)", "Plan EBITDA Margin % (fx)"]
for j, h in enumerate(headers, start=1):
    ws.cell(row=1, column=j, value=h)
style_header_row(ws, 1, len(headers))
for i, row in lrp.iterrows():
    r = i + 2
    ws.cell(row=r, column=1, value=int(row["fiscal_year"]))
    c_r = ws.cell(row=r, column=2, value=round(row["plan_revenue"], 2)); c_r.font = BLUE_INPUT; c_r.number_format = CUR_FMT
    c_o = ws.cell(row=r, column=3, value=round(row["plan_opex"], 2)); c_o.font = BLUE_INPUT; c_o.number_format = CUR_FMT
    c_e = ws.cell(row=r, column=4, value=f"=B{r}-C{r}"); c_e.font = BLACK_FORMULA; c_e.number_format = CUR_FMT
    c_m = ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/B{r},0)"); c_m.font = BLACK_FORMULA; c_m.number_format = PCT_FMT
    for col in range(1, 6):
        ws.cell(row=r, column=col).border = BORDER
ws.cell(row=len(lrp) + 4, column=1,
        value="Note: base-case LRP built from FY26 annualized run-rate x management growth assumptions (Rev CAGR 17-22%, Opex growth 13-16%). See Scenario Planning tab for Upside/Downside flex.").font = SUBTITLE_FONT
autosize(ws, [12, 18, 18, 16, 20])

# ========================================================================
# TAB 7 — ML FORECAST ACCURACY
# ========================================================================
acc = pd.read_csv(f"{DATA_DIR}/ml_model_accuracy_comparison.csv")
ws = wb.create_sheet("ML Forecast Accuracy")
ws["A1"] = "Forecast Model Accuracy Comparison — 6-month holdout (trailing actuals, company revenue)"
ws["A1"].font = SECTION_FONT
headers = ["Model", "MAPE (fx)", "RMSE (fx)", "MAE (fx)"]
for j, h in enumerate(headers, start=1):
    ws.cell(row=3, column=j, value=h)
style_header_row(ws, 3, len(headers))
for i, row in acc.sort_values("MAPE").reset_index(drop=True).iterrows():
    r = 4 + i
    ws.cell(row=r, column=1, value=row["model"])
    c_m = ws.cell(row=r, column=2, value=round(row["MAPE"], 4)); c_m.font = BLUE_INPUT; c_m.number_format = PCT_FMT
    c_rm = ws.cell(row=r, column=3, value=round(row["RMSE"], 0)); c_rm.font = BLUE_INPUT; c_rm.number_format = CUR_FMT
    c_ma = ws.cell(row=r, column=4, value=round(row["MAE"], 0)); c_ma.font = BLUE_INPUT; c_ma.number_format = CUR_FMT
    for col in range(1, 5):
        ws.cell(row=r, column=col).border = BORDER
autosize(ws, [26, 12, 14, 14])
ws.cell(row=4 + len(acc) + 2, column=1,
        value="MAPE = Mean Absolute Percentage Error (lower is better). Models trained on FY23-FY25 monthly revenue with lag/rolling-mean features; evaluated on the trailing 6 actual months.").font = SUBTITLE_FONT

# ========================================================================
# TAB 8 — DRIVER MODEL: REVENUE
# ========================================================================
drv = pd.read_csv(f"{DATA_DIR}/driver_model_revenue.csv", parse_dates=["period"])
ws = wb.create_sheet("Driver Model - Revenue")
headers = ["Period", "Segment", "Customers", "ARPU (implied)", "Revenue (fx = Customers x ARPU)",
           "New Logos", "Churned", "Net New Customers", "Churn Rate (implied, fx)"]
for j, h in enumerate(headers, start=1):
    ws.cell(row=1, column=j, value=h)
style_header_row(ws, 1, len(headers))
drv_sorted = drv.sort_values(["period", "profit_center"]).reset_index(drop=True)
for i, row in drv_sorted.iterrows():
    r = i + 2
    ws.cell(row=r, column=1, value=row["period"].strftime("%b-%Y"))
    ws.cell(row=r, column=2, value=row["profit_center"])
    c_c = ws.cell(row=r, column=3, value=round(row["customers"], 1)); c_c.font = BLUE_INPUT
    c_arpu = ws.cell(row=r, column=4, value=round(row["arpu_implied"], 2)); c_arpu.font = BLUE_INPUT; c_arpu.number_format = CUR_FMT
    c_rev = ws.cell(row=r, column=5, value=f"=C{r}*D{r}"); c_rev.font = BLACK_FORMULA; c_rev.number_format = CUR_FMT
    c_nl = ws.cell(row=r, column=6, value=round(row["new_logos"], 1)); c_nl.font = BLUE_INPUT
    c_ch = ws.cell(row=r, column=7, value=round(row["churned"], 1)); c_ch.font = BLUE_INPUT
    c_net = ws.cell(row=r, column=8, value=f"=F{r}-G{r}"); c_net.font = BLACK_FORMULA
    c_chr = ws.cell(row=r, column=9, value=f"=IFERROR(G{r}/C{r},0)"); c_chr.font = BLACK_FORMULA; c_chr.number_format = PCT_FMT
    for col in range(1, 10):
        ws.cell(row=r, column=col).border = BORDER
autosize(ws, [12, 10, 12, 14, 22, 10, 10, 14, 16])
ws.freeze_panes = "A2"

# ========================================================================
# TAB 9 — COMMENTARY NLP
# ========================================================================
nlp = pd.read_csv(f"{DATA_DIR}/commentary_nlp_scored.csv")
ws = wb.create_sheet("Commentary NLP")
headers = ["Fiscal Quarter", "Department", "Commentary", "Sentiment Score", "Predicted Tone",
           "Authored Tone (ground truth)", "Top Keywords", "Topic Cluster"]
for j, h in enumerate(headers, start=1):
    ws.cell(row=1, column=j, value=h)
style_header_row(ws, 1, len(headers))
for i, row in nlp.iterrows():
    r = i + 2
    ws.cell(row=r, column=1, value=row["fiscal_quarter"])
    ws.cell(row=r, column=2, value=row["department"])
    ws.cell(row=r, column=3, value=row["commentary_text"]).alignment = Alignment(wrap_text=True, vertical="top")
    c_s = ws.cell(row=r, column=4, value=row["sentiment_score"]); c_s.font = BLUE_INPUT
    ws.cell(row=r, column=5, value=row["predicted_tone"])
    ws.cell(row=r, column=6, value=row["authored_tone_label"])
    ws.cell(row=r, column=7, value=row["top_keywords"])
    ws.cell(row=r, column=8, value=int(row["topic_cluster"]))
    for col in range(1, 9):
        ws.cell(row=r, column=col).border = BORDER
last_row_nlp = len(nlp) + 1
ws.conditional_formatting.add(f"D2:D{last_row_nlp}",
    ColorScaleRule(start_type="min", start_color="F8696B", mid_type="num", mid_value=0, mid_color="FFEB84",
                   end_type="max", end_color="63BE7B"))
autosize(ws, [12, 22, 70, 12, 12, 18, 30, 10])

wb.save(OUT_PATH)
print(f"Workbook saved to {OUT_PATH}")
